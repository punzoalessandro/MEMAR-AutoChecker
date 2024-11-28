#Created by Alessandro Punzo (https://github.com/punzoalessandro)

from queue import Queue
import re
import threading
import openpyxl
import os
import shutil
import datetime
import sys
import time
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
import tkinter as tk
from tkinter import messagebox
from openpyxl.utils.exceptions import InvalidFileException



oggi = datetime.date.today()  
home_dir = os.path.expanduser("~") 
download_dir = os.path.join(home_dir, "Downloads") 
file_pathzip = os.path.join(download_dir, "TOTALI_MEMAR.TXT.zip") 
file_pathexcel = os.path.join(download_dir, oggi.strftime(f"%d%m%Y")+".xlsx") 
message_queue = Queue() 
flag = True 



class MyHandler(FileSystemEventHandler):    #Classe che viene richiamata quando l'observer rileva nuovi file all'interno della cartella download
                                            #il metodo al suo interno contiene tutti i controlli e l'inserimento in coda dei messaggi da mostrare all'utente    
    def on_created(self, event):
        global flag #Flag per verificare che il controllo avvenga solo quando è la prima volta che quest'ultimo è stato fatto

        if os.path.exists(file_pathzip) and os.path.exists(file_pathexcel) and flag == True:    #Controlliamo che esistano i due file (File zip e excel) e che il controllo non sia stato ancora fatto

            mtime_dt = datetime.datetime.fromtimestamp(os.path.getmtime(file_pathzip))
            data_formattata = mtime_dt.strftime("%d/%m/%Y")
            oggi_formattato = oggi.strftime("%d/%m/%Y")

            if data_formattata == oggi_formattato:      #Se la data di modifica del file zip non è uguale alla data di oggi mostra all'utente un messaggio di errore
                shutil.unpack_archive(file_pathzip, download_dir)  #Se la data è corretta estrae il contenuto del file zip
                message_queue.put(f"I file sono stati estratti con successo in: {download_dir}")
            else:
                message_queue.put(f"L'archivio ZIP {file_pathzip} non esiste o è vecchio.")
                sys.exit()

            try:    #Apre il file txt estratto
                f = open(download_dir + "/TOTALI_MEMAR.txt")
            except FileNotFoundError:
                message_queue.put("Il file TOTALI_MEMAR.txt non è stato trovato.")
                sys.exit()

            try:    #Apre il file excel
                fexcel = openpyxl.load_workbook(file_pathexcel)
            except FileNotFoundError:
                message_queue.put("Il file Excel non è stato trovato.")
                sys.exit()
            except InvalidFileException:
                message_queue.put("Il file specificato non è un file Excel valido.")
                sys.exit()
            except PermissionError:
                message_queue.put("Non hai i permessi necessari per aprire il file Excel.")
                sys.exit()

            listabi = [
            "03054", "03075", "03084", "03087", "03119", "03127", "03136", "03141", "03157", "03158", 
            "03169", "03170", "03191", "03197", "03209", "03242", "03251", "03258", "03263", "03268", 
            "03279", "03281", "03301", "03310", "03318", "03323", "03330", "03338", "03345", "03361", 
            "03367", "03381", "03395", "03399", "03403", "03417", "03425", "03426", "03433", "03435", 
            "03441", "03475", "03492", "03589", "03640", "03660", "05000", "05009", "05029", "05036", 
            "05044", "05060", "05080", "05083", "05104", "05141", "05142", "05144", "05156", "05180", 
            "05234", "05262", "05296", "05297", "05360", "05385", "05387", "05398", "05420", "05452", 
            "05484", "05544", "05572", "05579", "05597", "05602", "05640", "05650", "05652", "05664", 
            "05747", "05764", "05772", "05776", "05786", "05792", "05793", "05856", "05875", "06015", 
            "06050", "06055", "06060", "06120", "06140", "06150", "06195", "06200", "06205", "06230", 
            "06245", "06255", "06270", "06285", "06295", "06300", "06375", "06045", "07067", "07074", 
            "07084", "07888", "08883", "08888", "08917", "09036", "09050", "09051", "09052", "09075", 
            "09087", "09258", "09980", "09981", "09982", "09983", "09990", "09991", "09992", "09994", 
            "09995", "09996", "09997", "09998", "09999", "10008", "10312", "11111", "11112", "12933", 
            "13258", "15107", "15122", "15158", "15227", "15409", "16026", "16030", "16033", "16265", 
            "16460", "16714", "19060", "19248", "19275", "19351", "19493", "20039", "31108", "31333", 
            "31369", "31934", "32072", "32372", "32377", "32462", "32489", "32545", "32652", "32653", 
            "32698", "32728", "32741", "32753", "32774", "32814", "32863", "32905", "32939", "32987", 
            "33089", "33142", "33154", "33167", "33188", "33276", "33326", "33351", "33396", "33397", 
            "33435", "33477", "33591", "33656", "33661", "33679", "33691", "35009", "35199", "35200", 
            "35253", "35275", "35318", "35334", "35391", "35395", "35448", "35463", "35490", "35618", 
            "35669", "35978", "36000", "36092", "41467", "43435", "53435", "63435", "73435", "76060", 
            "76245", "83075", "83435", "85000", "86060", "86245", "89006", "89007", "90001", "91934", 
            "92741", "93025", "93043", "93075", "93142", "93158", "93209", "93242", "93301", "93425", 
            "93435", "93591", "93661", "95000", "95036", "95080", "95107", "95262", "95296", "95484", 
            "95652", "96245", "96270", "96285", "99006", "99204", "99933"
            ]
            #lista degli abi che ci serve per verificare che quelli all'interno siano corretti e a popolare l'array che verrà confrontato con il file excel
            rows = len(listabi)
            cols = 3
            tupla_temp = [[0] * cols for _ in range(rows)]

            row_index = 0
            abi_presenti = 0

            for line in f: #Verifichiamo che l'abi sia presente nella lista e lo inseriamo nel primo campo dell'array 
                match = re.search(r'\b\d{4,5}\b', line)
                if match:
                    linea = match.group()
                    if len(linea) == 4:
                        linea_con_zero = "0" + linea
                        if linea_con_zero in listabi:
                            tupla_temp[row_index][0] = linea
                            row_index += 1
                            abi_presenti += 1
                    elif len(linea) == 5:
                        if linea in listabi:
                            tupla_temp[row_index][0] = linea[1:]
                            row_index += 1
                            abi_presenti += 1
            row_index = 0
            f.seek(0)

            for line in f:  
                if 'EUR' in line: #Per ricavare i totali dal file txt andremo a ricercare la parola 'EUR' dato che è presente sempre nella riga dove si trovano i due risultati
                    diviso = line.split('EUR', 1)
                    primaparte = diviso[0].strip()   #Estraiamo il primo totale
                    secondaparte = diviso[1].strip() #Estraiamo il secondo totale
                    totale1 = re.sub(r'\D', '', primaparte) 
                    totale2 = re.sub(r'\D', '', secondaparte)
                    tupla_temp[row_index][1] = totale1.rstrip('0') #
                                                                   # Inseriamo i due totali all'interno dell'array
                    tupla_temp[row_index][2] = totale2.rstrip('0') #
                    
                    row_index += 1

            index = 0
            tupla = []

            while index < abi_presenti: #Inserisco ogni array nell'array finale che contiene i risultati provenienti dal file txt
                tupla.append(tupla_temp[index]) 
                index += 1

            #Eseguo tutti i controlli riguardanti il file excel

            foglio = fexcel['Sheet1']
            indexcel = 3
            countergiusti = 0

            for riga in foglio.iter_rows(min_row=indexcel, values_only=True):      #Controllo che non ci siano valori nulli  
                if riga[3] != '' and riga[3] is not None:  
                    countergiusti += 1

            if countergiusti == abi_presenti: #Confronto il numero di valori corretti del file excel con il numero di abi presenti nel file txt
                message_queue.put("Nessun problema nel file excel")
            elif countergiusti == 0:
                message_queue.put("File Excel vuoto, controllare e contattare Caricese")    
            else:
                message_queue.put("Le righe non corrispondono con il file txt, controllare")

            arrexcel = [[0] * 3 for _ in range(countergiusti)]
            indxrw = 0

            #Nota: questa parte poteva essere messa all'interno del controllo commentato in precedenza    

            for riga in foglio.iter_rows(min_row=3, values_only=True,max_row=countergiusti+2): #Inserisco ogni riga in un array suddiviso come 'abi,totale1,totale2"
                arrexcel[indxrw][0]=str(riga[0]).lstrip('0')
                arrexcel[indxrw][1]=str(riga[2]).replace(".","").rstrip('0')
                arrexcel[indxrw][2]=str(riga[3]).replace(".","").rstrip('0')
                indxrw += 1

            #Ordino i due array
            array1_sorted = sorted(arrexcel)    
            array2_sorted = sorted(tupla)
            

            #Se gli array sono uguali vuol dire che tutti i valori corrispondono
            if array1_sorted == array2_sorted:
                message_queue.put("I due file sono uguali")
            else:
                message_queue.put("I due file non sono uguali, controlla")    

            flag=False


def show_message(): #Funzione che serve per mostrare i messaggi messi nella coda
    
    while not message_queue.empty():
        message = message_queue.get()
        messagebox.showwarning("Controllo Memar", message)


    root.after(100, show_message)

def start_watchdog(): #Funzione che serve per avviare l'osservatore che monitora l'arrivo di nuovi file nella cartella download

    event_handler = MyHandler()
    observer = Observer()
    observer.schedule(event_handler, download_dir, recursive=False)
    observer.start()
    try:
        while True:
            time.sleep(1) 
    except KeyboardInterrupt:
        observer.stop()
    observer.join()

if __name__ == "__main__": #Funzione main

    root = tk.Tk()
    root.withdraw()


    watchdog_thread = threading.Thread(target=start_watchdog)
    watchdog_thread.daemon = True  # Termina il thread quando l'applicazione si chiude
    watchdog_thread.start()

    # Avvia la funzione che mostra i messaggi nella coda
    root.after(100, show_message)

    # Avvia il mainloop di Tkinter nel thread principale
    root.mainloop()


